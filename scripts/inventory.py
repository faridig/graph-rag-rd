#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inventory an experiment workbook so you can SEE its full structure before extracting.

Usage:
    python inventory.py FILE.xlsx              # values, every non-empty row, with row numbers
    python inventory.py FILE.xlsx --formulas   # also print formula strings where present
    python inventory.py FILE.csv               # CSV files are dumped too

Prints, per sheet: dimensions, then each non-empty row as 'rownum || c1 | c2 | ...'.
This is the map you read to locate metadata, per-run blocks, conditions and results.
Do NOT copy numbers from this dump into your JSON — read them from the cells in code.
"""
import sys, os, csv

def dump_csv(path):
    print(f"\n############ CSV: {os.path.basename(path)}")
    with open(path, newline='', encoding='utf-8', errors='replace') as f:
        try:
            dialect = csv.Sniffer().sniff(f.read(4096)); f.seek(0)
        except Exception:
            dialect = csv.excel; f.seek(0)
        for i, row in enumerate(csv.reader(f, dialect), 1):
            if any(c.strip() for c in row):
                print(i, "||", " | ".join(row))

def dump_xlsx(path, show_formulas):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path, data_only=False) if show_formulas else None
    print(f"############ WORKBOOK: {os.path.basename(path)}")
    print("SHEETS:", wb.sheetnames)
    for ws in wb.worksheets:
        print(f"\n=== SHEET: {ws.title}  (max_row={ws.max_row}, max_col={ws.max_column}) ===")
        wsf = wbf[ws.title] if wbf else None
        for r in range(1, ws.max_row + 1):
            cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(c is not None and str(c).strip() != "" for c in cells):
                continue
            line = " | ".join("" if c is None else str(c) for c in cells)
            print(f"{r} || {line}")
            if wsf:
                fcells = []
                for c in range(1, ws.max_column + 1):
                    fv = wsf.cell(r, c).value
                    if isinstance(fv, str) and fv.startswith("="):
                        fcells.append(f"{openpyxl.utils.get_column_letter(c)}{r}={fv}")
                if fcells:
                    print("     [formulas] " + " ; ".join(fcells))

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    path = sys.argv[1]
    show_formulas = "--formulas" in sys.argv[2:]
    if not os.path.exists(path):
        print("File not found:", path); sys.exit(1)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        dump_xlsx(path, show_formulas)
    elif ext in (".csv", ".tsv"):
        dump_csv(path)
    else:
        print("Unsupported extension:", ext, "- expected .xlsx/.xlsm/.csv/.tsv"); sys.exit(1)

if __name__ == "__main__":
    main()
