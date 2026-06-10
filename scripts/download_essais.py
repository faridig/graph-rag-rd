#!/usr/bin/env python3
"""
Télécharge tous les fichiers liés dans la colonne 'Lien vers essai' du RÉPERTOIRE Excel.
Authentification via MSAL device code flow (navigateur — aucune config admin requise).

Usage:
    pip install msal requests
    python scripts/download_essais.py
    python scripts/download_essais.py --dest data/essais --sheet "Répertoire"
"""

import argparse
import base64
import io
import logging
import re
import sys
from pathlib import Path

import openpyxl

import msal
import requests

# ── Constantes ───────────────────────────────────────────────────────────────

# Azure CLI client ID — app Microsoft first-party, pré-approuvée dans la plupart des tenants.
# Si le tenant bloque même ce client, voir commentaire en bas de fichier.
CLIENT_ID = "e5e5bc3b-aeb0-4fbb-8cd8-76a90a8fcf11"
TENANT_ID = "21aa992f-60cb-4048-b8b9-64e1fb98e11f"
SCOPES = [
    "https://graph.microsoft.com/Files.Read.All",
    "https://graph.microsoft.com/Sites.Read.All",
]

SITE_HOSTNAME = "nxtfoodfr.sharepoint.com"
SITE_PATH = "/sites/RD"
# ID extrait du sourcedoc= de l'URL SharePoint
REPERTOIRE_ITEM_ID = "3E7753D0-1871-4786-9EF4-040E3B08AF38"

GRAPH = "https://graph.microsoft.com/v1.0"

# ── Auth ──────────────────────────────────────────────────────────────────────


def get_token() -> str:
    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        token_cache=msal.SerializableTokenCache(),
    )
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            return result["access_token"]

    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(f"Impossible d'initier device flow : {flow}")

    print(f"\n{'='*60}")
    print(flow["message"])
    print(f"{'='*60}\n")

    result = app.acquire_token_by_device_flow(flow)
    if "access_token" not in result:
        raise RuntimeError(
            f"Authentification échouée : {result.get('error_description', result)}\n\n"
            "Si l'erreur mentionne 'admin consent', demandez à votre admin IT d'approuver\n"
            "l'application Azure CLI (client_id 04b07795-...) pour le tenant nxtfood.fr,\n"
            "ou enregistrez votre propre app dans Azure AD (voir README pour la procédure)."
        )
    return result["access_token"]


# ── Helpers Graph API ─────────────────────────────────────────────────────────


def graph_get(session: requests.Session, path: str, **kwargs) -> dict:
    resp = session.get(f"{GRAPH}{path}", **kwargs)
    resp.raise_for_status()
    return resp.json()


def get_drive_id(session: requests.Session) -> str:
    """Retourne le driveId du document library par défaut du site RD."""
    return _get_drive_id_for_site(session, SITE_PATH)


def _get_drive_id_for_site(session: requests.Session, site_path: str) -> str:
    """Retourne le driveId du document library par défaut pour un site SharePoint."""
    site = graph_get(session, f"/sites/{SITE_HOSTNAME}:{site_path}")
    site_id = site["id"]
    drives = graph_get(session, f"/sites/{site_id}/drives", params={"$select": "id,name,driveType"})
    for drive in drives["value"]:
        if drive.get("driveType") == "documentLibrary":
            return drive["id"]
    return drives["value"][0]["id"]


# Cache des drive IDs par site pour éviter les appels répétés
_drive_id_cache: dict[str, str] = {}


# ── Lecture Excel ─────────────────────────────────────────────────────────────


def _extract_hyperlink_url(formula: str) -> str | None:
    """Extrait l'URL d'une formule =HYPERLINK("url", ...) ou =LIEN_HYPERTEXTE(...)."""
    m = re.match(r'=(?:HYPERLINK|LIEN_HYPERTEXTE)\("([^"]+)"', formula, re.IGNORECASE)
    return m.group(1) if m else None


def _is_sharepoint_url(text: str) -> bool:
    return "sharepoint.com" in text or "nxtfoodfr" in text


def _sanitize_filename(name: str) -> str:
    """Remplace les caractères interdits dans un nom de fichier Linux."""
    return re.sub(r'[/\x00]', '_', name).strip()


def _normalize_sharepoint_url(url: str) -> str:
    """
    Normalise les URLs SharePoint qui contiennent des /../ résiduels.
    Ex: https://nxtfoodfr.sharepoint.com/sites/RD/.../../../../../:x:/s/RD/XXX
    → https://nxtfoodfr.sharepoint.com/:x:/s/RD/XXX
    """
    # Cherche le pattern /:x:/ ou /:w:/ ou /:b:/ (liens de partage SharePoint)
    m = re.search(r'(https://[^/]+)((?:/[^/]+)*/\.\.(?:/\.\.)*/)?(:[a-z]:/s/.+)', url)
    if m:
        return m.group(1) + "/" + m.group(3)
    return url


def get_links_from_excel(
    session: requests.Session,
    drive_id: str,
    target_sheet: str | None = None,
) -> list[tuple[str, str]]:
    """
    Télécharge le fichier Excel et lit les hyperliens natifs de la colonne K via openpyxl.
    Retourne une liste de (label, url).
    """
    # 1. Télécharger le fichier Excel en mémoire
    print("Téléchargement du fichier Excel...")
    resp = session.get(
        f"{GRAPH}/drives/{drive_id}/items/{REPERTOIRE_ITEM_ID}/content",
        allow_redirects=True,
    )
    resp.raise_for_status()
    print(f"Fichier téléchargé ({len(resp.content) // 1024} KB)")

    # 2. Ouvrir avec openpyxl (keep_vba=False suffit pour les hyperliens)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    print(f"Feuilles disponibles : {wb.sheetnames}")
    sheet_name = target_sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable. Disponibles : {wb.sheetnames}")
    ws = wb[sheet_name]
    print(f"Feuille : {sheet_name} ({ws.max_row} lignes × {ws.max_column} colonnes)")

    # 3. Colonne K = index 11 (1-based dans openpyxl)
    COL_K = 11
    links: list[tuple[str, str]] = []
    skipped_empty = 0
    skipped_external = 0

    for row in ws.iter_rows(min_row=1, max_col=COL_K, min_col=COL_K):
        cell = row[0]
        if cell.hyperlink is None:
            skipped_empty += 1
            continue
        url: str = str(cell.hyperlink.target).strip()
        label: str = _sanitize_filename(str(cell.value).strip() if cell.value else "")
        if not url:
            skipped_empty += 1
            continue
        # Ignorer uniquement les ancres et mailto
        if url.startswith("#") or url.startswith("mailto:"):
            skipped_external += 1
            continue
        url = _normalize_sharepoint_url(url)
        print(f"  row {cell.row:3d} | {label[:35]:35s} | {url[:70]}")
        links.append((label, url))

    print(
        f"{len(links)} hyperliens trouvés en colonne K "
        f"({skipped_empty} vides, {skipped_external} ignorés)"
    )
    return links


# ── Résolution et téléchargement ──────────────────────────────────────────────


def _encode_sharing_url(url: str) -> str:
    """Encode une URL pour l'endpoint Graph /shares/."""
    b64 = base64.urlsafe_b64encode(url.encode()).decode().rstrip("=")
    return f"u!{b64}"


SITE_BASE = "https://nxtfoodfr.sharepoint.com"


def _strip_leading_dotdot(url: str) -> str:
    """Retire les préfixes ../ et retourne (nb_niveaux, chemin_propre)."""
    clean = re.sub(r'^(\.\./)+', '', url)
    return clean


def resolve_sharepoint_link(
    session: requests.Session, url: str, drive_id: str | None = None
) -> dict | None:
    """Résout un lien en driveItem Graph. Gère URLs absolues, relatives et sourcedoc GUIDs."""

    if url.startswith("file:///"):
        print(f"  ✗ Chemin local Windows (inaccessible) : {url[:80]}")
        return None

    # Résolution des chemins relatifs
    is_relative = url.startswith("..") or url.startswith("./")
    if is_relative:
        clean = _strip_leading_dotdot(url)

        # Cas A : lien de partage (:x:, :w:, :b: …)
        if clean.startswith(":"):
            url = f"{SITE_BASE}/{clean}"
            is_relative = False

        # Cas B : chemin absolu dans le site (sites/RD/...)
        elif clean.startswith("sites/"):
            url = f"{SITE_BASE}/{clean}"
            is_relative = False

        # Cas C : chemin relatif dans le drive (2. Extrusion/...) → strategy drive:root
        # On garde is_relative=True pour utiliser la stratégie /root:/ ci-dessous

    # Stratégie 1 : GUID sourcedoc → accès direct
    if not is_relative:
        m = re.search(r'sourcedoc=%7[Bb]([A-Fa-f0-9-]+)%7[Dd]', url)
        if m and drive_id:
            item_id = m.group(1)
            try:
                return graph_get(session, f"/drives/{drive_id}/items/{item_id}")
            except requests.HTTPError:
                pass

    # Stratégie 2 : chemin relatif dans la bibliothèque → /drives/{id}/root:/{path}
    if is_relative and drive_id:
        from urllib.parse import unquote
        clean = _strip_leading_dotdot(url)
        clean = clean.split("?")[0]
        path = unquote(clean).strip("/")

        # Détection chemin cross-site : "SiteName/Documents partages/..."
        # Ex: "RDIndustrieCollaboration/Documents partages/Essais industriels/..."
        cross_site_match = re.match(r'^([A-Za-z0-9_-]+)/([^/]+)/(.+)$', path)
        if cross_site_match:
            site_name = cross_site_match.group(1)
            file_path = cross_site_match.group(3)   # chemin sous la bibliothèque (ignore lib_name)
            cross_site_path = f"/sites/{site_name}"
            if cross_site_path != SITE_PATH:
                # Résoudre avec le drive du site cible
                cross_drive_id = _drive_id_cache.get(cross_site_path)
                if cross_drive_id is None:
                    try:
                        cross_drive_id = _get_drive_id_for_site(session, cross_site_path)
                        _drive_id_cache[cross_site_path] = cross_drive_id
                    except requests.HTTPError as exc_site:
                        st = exc_site.response.status_code if exc_site.response is not None else "?"
                        print(f"  ✗ Site {site_name} inaccessible ({st})")
                        cross_drive_id = None
                if cross_drive_id:
                    try:
                        return graph_get(session, f"/drives/{cross_drive_id}/root:/{file_path}")
                    except requests.HTTPError as exc_cs:
                        st = exc_cs.response.status_code if exc_cs.response is not None else "?"
                        print(f"  ✗ cross-site drive/root:/ ({st}) : {file_path[:70]}")

        try:
            return graph_get(session, f"/drives/{drive_id}/root:/{path}")
        except requests.HTTPError as exc2:
            status2 = exc2.response.status_code if exc2.response is not None else "?"
            print(f"  ✗ drive/root:/ ({status2}) : {path[:70]}")

    # Stratégie 3 : lien de partage via /shares/
    if not is_relative:
        encoded = _encode_sharing_url(url)
        try:
            return graph_get(session, f"/shares/{encoded}/driveItem")
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "?"
            print(f"  ✗ Résolution échouée ({status}) : {url[:80]}")

    return None


def download_item(session: requests.Session, item: dict, dest_dir: Path) -> bool:
    """Télécharge un driveItem dans dest_dir. Retourne True si succès."""
    name: str = item["name"]
    dest = dest_dir / name

    if dest.exists():
        print(f"  ~ déjà présent : {name}")
        return True

    download_url: str | None = item.get("@microsoft.graph.downloadUrl")
    if download_url:
        resp = session.get(download_url)
    else:
        drive_id = item["parentReference"]["driveId"]
        item_id = item["id"]
        resp = session.get(
            f"{GRAPH}/drives/{drive_id}/items/{item_id}/content",
            allow_redirects=True,
        )

    try:
        resp.raise_for_status()
    except requests.HTTPError as exc:
        print(f"  ✗ Téléchargement échoué ({exc.response.status_code}) : {name}")
        return False

    dest.write_bytes(resp.content)
    size_kb = len(resp.content) // 1024
    print(f"  ✓ {name} ({size_kb} KB)")
    return True


# ── Main ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(description="Télécharge les fichiers 'Lien vers essai' depuis SharePoint")
    parser.add_argument("--dest", default="data/essais", help="Dossier de destination (défaut: data/essais)")
    parser.add_argument("--sheet", default=None, help="Nom de la feuille Excel (défaut: première feuille)")
    args = parser.parse_args()

    dest_dir = Path(args.dest)
    dest_dir.mkdir(parents=True, exist_ok=True)
    log_path = dest_dir / "download.log"

    # Logging : console + fichier
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_path, encoding="utf-8"),
        ],
    )
    log = logging.getLogger("download")

    log.info("Dossier de destination : %s", dest_dir.resolve())
    log.info("Log : %s", log_path.resolve())

    log.info("[1/4] Authentification Microsoft...")
    token = get_token()
    session = requests.Session()
    session.headers["Authorization"] = f"Bearer {token}"

    log.info("[2/4] Connexion au site SharePoint...")
    drive_id = get_drive_id(session)
    log.info("Drive ID : %s", drive_id)

    log.info("[3/4] Lecture du fichier RÉPERTOIRE Excel...")
    links = get_links_from_excel(session, drive_id, target_sheet=args.sheet)

    if not links:
        log.error("Aucun lien trouvé. Vérifiez le nom de la colonne dans l'Excel.")
        sys.exit(1)

    log.info("[4/4] Téléchargement de %d fichiers...", len(links))
    ok = 0
    skipped = 0
    failed: list[tuple[str, str, str]] = []

    for i, (label, url) in enumerate(links, 1):
        log.info("[%d/%d] %s", i, len(links), label or url[:80])
        item = resolve_sharepoint_link(session, url, drive_id=drive_id)
        if item is None:
            log.warning("LIEN MORT : label=%r url=%s", label, url)
            failed.append((label, url, "lien mort / résolution impossible"))
            continue

        # Nom = label de la colonne K + extension du fichier SharePoint (si pas déjà présente)
        sharepoint_name: str = item["name"]
        ext = Path(sharepoint_name).suffix  # ex: ".xlsx"
        if label and not label.lower().endswith(ext.lower()):
            filename = label + ext
        else:
            filename = label or sharepoint_name
        dest = dest_dir / filename

        # Always log the real webUrl from Graph — useful for sharepoint_urls.py even if already downloaded
        web_url = item.get("webUrl")
        if web_url and label:
            log.info("WEBURL: label=%r url=%s", label, web_url)

        if dest.exists():
            log.info("  ~ déjà présent : %s", filename)
            skipped += 1
            continue

        download_url: str | None = item.get("@microsoft.graph.downloadUrl")
        if download_url:
            resp = session.get(download_url)
        else:
            did = item["parentReference"]["driveId"]
            iid = item["id"]
            resp = session.get(f"{GRAPH}/drives/{did}/items/{iid}/content", allow_redirects=True)

        try:
            resp.raise_for_status()
            dest.write_bytes(resp.content)
            log.info("  ✓ %s (%d KB)", filename, len(resp.content) // 1024)
            ok += 1
        except requests.HTTPError as exc:
            log.error("  ✗ ÉCHEC TÉLÉCHARGEMENT (%s) : %s — %s", exc.response.status_code, filename, url)
            failed.append((label, url, f"HTTP {exc.response.status_code}"))

    log.info("=" * 60)
    log.info("Terminé : %d téléchargés, %d déjà présents, %d échecs", ok, skipped, len(failed))
    if failed:
        log.warning("ÉCHECS (%d) :", len(failed))
        for label, url, reason in failed:
            log.warning("  [%s] %r — %s", reason, label, url)
    log.info("Fichiers dans : %s", dest_dir.resolve())


if __name__ == "__main__":
    main()

# ── Note : si l'erreur "admin consent" persiste ───────────────────────────────
# L'Azure CLI client ID est normalement pré-approuvé. Si le tenant nxtfood.fr
# a désactivé ce client, deux options :
#
# Option A — Enregistrer votre propre app Azure AD (5 min) :
#   1. portal.azure.com → Azure Active Directory → App registrations → New
#   2. Redirect URI : http://localhost (Public client)
#   3. API permissions : Files.Read.All + Sites.Read.All (Delegated)
#   4. Remplacer CLIENT_ID ci-dessus par votre Application (client) ID
#   5. Demander à l'admin de "Grant admin consent" pour ces seules permissions
#
# Option B — Utiliser les cookies de navigateur (sans app Azure) :
#   Voir scripts/download_essais_cookies.py (non fourni — demander si nécessaire)
